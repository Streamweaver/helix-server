import typing
import json
from datetime import datetime
from django.core.serializers.json import DjangoJSONEncoder
from django.utils import timezone
from rest_framework import viewsets
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from rest_framework.permissions import AllowAny
from rest_framework.decorators import action
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import mixins
from drf_spectacular.utils import extend_schema
from django.db import models
from django.db.models import (
    F, Case, When, Q
)

from apps.contrib.commons import DATE_ACCURACY
from apps.country.models import Country
from apps.entry.models import ExternalApiDump, Figure, OSMName
from apps.common.utils import (
    EXTERNAL_ARRAY_SEPARATOR,
    EXTERNAL_FIELD_SEPARATOR,
)
from apps.event.models import EventCode
from apps.crisis.models import Crisis

from .models import (
    Conflict, Disaster, DisplacementData, GiddFigure, IdpsSaddEstimate,
    StatusLog, PublicFigureAnalysis
)
from .serializers import (
    CountrySerializer,
    ConflictSerializer,
    DisaggregationSerializer,
    DisasterSerializer,
    DisplacementDataSerializer,
    PublicFigureAnalysisSerializer,
)
from .rest_filters import (
    DisaggregationFilterSet,
    RestConflictFilterSet,
    RestDisasterFilterSet,
    RestDisplacementDataFilterSet,
    IdpsSaddEstimateFilter,
    PublicFigureAnalysisFilterSet,
    DisaggregationPublicFigureAnalysisFilterSet,
)
from utils.common import track_gidd, client_id


def _get_location_accuracy_label(accuracy):
    if accuracy is None:
        return None
    return OSMName.OSM_ACCURACY.get(accuracy).label


def _get_location_type_label(type):
    if type is None:
        return None
    return OSMName.IDENTIFIER.get(type).label


def _get_event_code_label(key: str):
    if key is None:
        return None
    return EventCode.EVENT_CODE_TYPE.get(int(key)).label


def _get_location_accuracy_labels(
    location_accuracy: typing.List[typing.Tuple[int]]
) -> str:
    return string_join(
        EXTERNAL_ARRAY_SEPARATOR,
        [_get_location_accuracy_label(accuracy) for accuracy in location_accuracy]
    )


def _get_location_type_labels(
    location_type: typing.List[typing.Tuple[int]]
) -> str:
    return string_join(
        EXTERNAL_ARRAY_SEPARATOR,
        [_get_location_type_label(type) for type in location_type]
    )


def string_join(
    separator: str,
    data: typing.List[str],
) -> str:
    return separator.join([
        str(item)
        for item in data
        if item is not None
    ])


def remove_null_from_dict(data: dict) -> dict:
    return {
        key: value
        for key, value in data.items()
        if value is not None
    }


@client_id
class ListOnlyViewSetMixin(mixins.ListModelMixin, viewsets.GenericViewSet):
    def get(self, request, *args, **kwargs):
        return self.list(request, *args, **kwargs)


class CountryViewSet(ListOnlyViewSetMixin):
    serializer_class = CountrySerializer
    lookup_field = 'iso3'
    filter_backends = (DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter)
    filterset_fields = ['id']

    def get_queryset(self):
        track_gidd(
            self.request.GET.get('client_id'),
            ExternalApiDump.ExternalApiType.GIDD_COUNTRY_REST,
            viewset=self,
        )
        return Country.objects.all()


class ConflictViewSet(ListOnlyViewSetMixin):
    serializer_class = ConflictSerializer
    filterset_class = RestConflictFilterSet

    def get_queryset(self):
        track_gidd(
            self.request.GET.get('client_id'),
            ExternalApiDump.ExternalApiType.GIDD_CONFLICT_REST,
            viewset=self,
        )
        return Conflict.objects.all().select_related('country')


class DisasterViewSet(ListOnlyViewSetMixin):
    serializer_class = DisasterSerializer
    filterset_class = RestDisasterFilterSet

    def get_queryset(self):
        api_type = ExternalApiDump.ExternalApiType.GIDD_DISASTER_REST
        if self.action == 'export':
            api_type = ExternalApiDump.ExternalApiType.GIDD_DISASTER_EXPORT_REST

        track_gidd(
            self.request.GET.get('client_id'),
            api_type,
            viewset=self,
        )
        return Disaster.objects.select_related('country')

    @staticmethod
    def get_displacement_status(displacement_occurred: typing.List[int]) -> str:
        if not displacement_occurred:
            return ""
        elif Figure.DISPLACEMENT_OCCURRED.BEFORE.value in displacement_occurred:
            return "Displacement reporting preventive evacuations"
        return "Displacement without preventive evacuations reported"

    @extend_schema(responses=DisasterSerializer(many=True))
    @action(
        detail=False,
        methods=["get"],
        url_path="disaster-export",
        permission_classes=[AllowAny],
        pagination_class=None,
    )
    def export(self, request):
        """
        Export disaster
        """
        qs = self.filter_queryset(self.get_queryset())
        wb = Workbook(write_only=True)
        ws = wb.create_sheet('1_Disaster_Displacement_data')
        ws.append([
            'ISO3',
            'Country / Territory',
            'Year',
            'Event Name',
            'Date of Event (start)',
            'Disaster Internal Displacements',
            'Disaster Internal Displacements (Raw)',
            'Hazard Category',
            'Hazard Type',
            'Hazard Sub Type',
            'Event Codes (Code:Type)',
            'Event ID',
            'Displacement occurred',
        ])

        for disaster in qs:
            ws.append(
                [
                    disaster.country.iso3,
                    disaster.country.name,
                    disaster.year,
                    disaster.event_name,
                    disaster.start_date,
                    disaster.new_displacement_rounded,
                    disaster.new_displacement,
                    disaster.hazard_category_name,
                    disaster.hazard_type_name,
                    disaster.hazard_sub_type_name,
                    EXTERNAL_ARRAY_SEPARATOR.join(
                        [f"{key}{EXTERNAL_FIELD_SEPARATOR}{value}" for key, value in zip(
                            disaster.event_codes,
                            disaster.event_codes_type
                        )]
                    ),
                    disaster.event_id,
                    self.get_displacement_status(disaster.displacement_occurred),
                ]
            )

        ws2 = wb.create_sheet('README')
        readme_text = [
            ['TITLE: Global Internal Displacement Database (GIDD) - Disasters'],
            [],
            ['FILENAME: IDMC_GIDD_Disasters_Internal_Displacement_Data'],
            [],
            ['SOURCE: Internal Displacement Monitoring Centre (IDMC)'],
            [],
            [f'DATE EXTRACTED: {datetime.now().strftime("%B %d, %Y")}'],
            [],
            [f'LAST UPDATE: {StatusLog.last_release_date()}'],
            [],
            ['DESCRIPTION:'],
            [
                'The Internal Displacement Monitoring Centre (IDMC) monitors internal displacement events globally, '
                'triggered by disasters, conflict, and other forms of violence. It gathers and analyses both '
                'structured and unstructured secondary data from diverse sources—including government agencies, '
                'UN agencies, the International Federation of the Red Cross and Red Crescent, and the media.'
            ],
            [],
            [
                'IDMC analysts rigorously analyse and triangulate all reported data. The data undergo thorough quality '
                'control processes, involving engagement with primary data collectors for peer review and validation. '
                'This meticulous approach guarantees that the data reported by IDMC reflects high accuracy.'
            ],
            [],
            [
                'The data in the Global Internal Displacement Database (GIDD) is annually validated and peer-reviewed, '
                'having passed through various quality control processes in consultation with different UN agencies, '
                'goverments and local data providers.'
            ],
            [],
            [
                'The GIDD database documents displacement due to conflict from 2009 to 2023 and disaster-induced '
                'displacement from 2008 to 2023. For detailed definitions and more comprehensive descriptions, please '
                'refer to the IDMC Monitoring Tools (https://www.internal-displacement.org/monitoring-tools).'
            ],
            [],
            ['KEY DEFINITIONS:'],
            [],
            [
                'Internal Displacements (flows): This metric represents the number of internal displacements, or '
                'internal displacement population flows, reported from January 1st to December 31st of a reporting year. '
                'This figure may include individuals who are displaced multiple times during the year by different events.'
            ],
            [
                'Total number of Internally Displaced Persons (IDPs) (stocks): This metric represents the total number '
                'of people living in situations of internal displacement as of the end of the reporting year, '
                'specifically on December 31st of each year.'
            ],
            [
                'Disaster displacement: Refers to situations where people are forced to leave their homes or places of '
                'habitual residence as a result, or in anticipation of the negative impact of natural hazards.'
            ],
            [
                'Disaster: A serious disruption of the functioning of a community or a society involving widespread '
                'human, material, economic or environmental losses and impacts, which exceeds the ability of the '
                'affected community or society to cope using its own resources (UNSDR).'
            ],
            [],
            [],
            [
                'USE LICENSE: This content is licensed under CC BY-NC. Detailed licensing information is available at '
                'Creative Commons License (See: https://creativecommons.org/licenses/by-nc/4.0/).'
            ],
            [],
            [
                'COVERAGE: Global. The GIDD provides data on internal displacements triggered by disasters dates back '
                'to 2008, and the metrics on the total number of IDPs from disaster-related events are available from '
                '2019 onwards.'
            ],
            [],
            ['CITATION:'],
            [
                'All derived work from IDMC data could cite IDMC following this example: Internal Displacement '
                'Monitoring Centre. Global Internal Displacement Database - Disasters. IDMC (2023). Available at: '
                'https://www.internal-displacement.org/database/displacement-data/ (Accessed: [date of access]).'
            ],
            [],
            ['CONTACT: info@idmc.ch'],
            [],
        ]

        for item in readme_text:
            ws2.append(item)
        ws2.append([])
        ws2.append(['DATA DESCRIPTION: 1_Disaster_Displacement_data'])
        ws2.append([])

        table = [
            ['ISO3: Represents the ISO 3166-1 alpha-3 code. The code \'AB9\' is assigned to the Abyei Area.'],
            ['Country / Territory: Short name of the country or territory.'],
            ['Year: Indicates the year for which displacement data are reported.'],
            [
                'Event Name: Common or official event name for the event, if available. Otherwise, events are coded '
                'based on the country, type of hazard, location, and event start date.'
            ],
            ['Date of Event (Start): Approximate start date of the event.'],
            [
                'Disaster Internal Displacements: Total number of internal displacements reported (rounded figures at '
                'national level), as a result of disasters over the reporting year.Units are recorded as \'internal '
                'displacement flows\' or \'internal displacement movements.\' '
            ],
            [
                'Disaster Internal Displacements raw: Total number of internal displacements reported (not rounded), '
                'as a result of disasters over the reporting year. Units are recorded as \'internal displacement flows\' '
                'or \'internal displacement movements.\' '
            ],
            ['Hazard Category: Hazard category based on the CRED EM-DAT classification.'],
            ['Hazard Type: Hazard type as categorized by CRED EM-DAT.'],
            ['Hazard Sub-Type: Specific sub-type of the hazard based on CRED EM-DAT.'],
            [
                'Event Codes (Code:Type): Unique codes such as the GLIDE number and other database-specific codes used '
                'to identify and track specific events across various databases.'
            ],
            ['Event ID: Unique identifier for events as assigned by IDMC.'],
            [
                'Displacement Occurred: This field contains values that represent if preventive evacuations were reported. '
                'These evacuations are the result of existing early warning systems.'
            ],
        ]
        for item in table:
            ws2.append(item)
        response = HttpResponse(content=save_virtual_workbook(wb))
        filename = 'IDMC_GIDD_Disasters_Internal_Displacement_Data.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        response['Content-Type'] = 'application/octet-stream'
        return response


class DisplacementDataViewSet(ListOnlyViewSetMixin):
    serializer_class = DisplacementDataSerializer
    filterset_class = RestDisplacementDataFilterSet

    def get_queryset(self):
        api_type = ExternalApiDump.ExternalApiType.GIDD_DISPLACEMENT_REST
        if self.action == 'export':
            api_type = ExternalApiDump.ExternalApiType.GIDD_DISPLACEMENT_EXPORT_REST

        track_gidd(
            self.request.GET.get('client_id'),
            api_type,
            viewset=self,
        )
        return DisplacementData.objects.all()

    def export_conflicts(self, ws, qs):
        ws.append([
            'ISO3',
            'Name',
            'Year',
            'Conflict Stock Displacement',
            'Conflict Stock Displacement (Raw)',
            'Conflict Internal Displacements',
            'Conflict Internal Displacements (Raw)',
        ])
        for item in qs:
            ws.append([
                item.iso3,
                item.country_name,
                item.year,
                item.conflict_total_displacement_rounded,
                item.conflict_total_displacement,
                item.conflict_new_displacement_rounded,
                item.conflict_new_displacement,
            ])

    def export_disasters(self, ws, qs):
        ws.append([
            'ISO3',
            'Name',
            'Year',
            'Disaster Internal Displacements',
            'Disaster Internal Displacements (Raw)',
            'Disaster Stock Displacement',
            'Disaster Stock Displacement (Raw)'
        ])
        for item in qs:
            ws.append([
                item.iso3,
                item.country_name,
                item.year,
                item.disaster_new_displacement_rounded,
                item.disaster_new_displacement,
                item.disaster_total_displacement_rounded,
                item.disaster_total_displacement,
            ])

    def export_displacements(self, ws, qs):
        ws.append([
            'ISO3',
            'Name',
            'Year',
            'Conflict Stock Displacement',
            'Conflict Stock Displacement (Raw)',
            'Conflict Internal Displacements',
            'Conflict Internal Displacements (Raw)',
            'Disaster Internal Displacements',
            'Disaster Internal Displacements (Raw)',
            'Disaster Stock Displacement',
            'Disaster Stock Displacement (Raw)'
        ])
        for item in qs:
            ws.append([
                item.iso3,
                item.country_name,
                item.year,
                item.conflict_total_displacement_rounded,
                item.conflict_total_displacement,
                item.conflict_new_displacement_rounded,
                item.conflict_new_displacement,
                item.disaster_new_displacement_rounded,
                item.disaster_new_displacement,
                item.disaster_total_displacement_rounded,
                item.disaster_total_displacement,
            ])

    @extend_schema(responses=DisplacementDataSerializer(many=True))
    @action(
        detail=False,
        methods=["get"],
        url_path="displacement-export",
        permission_classes=[AllowAny],
        pagination_class=None,
    )
    def export(self, request):
        """
        Export displacements, conflict and disaster
        """

        # Track export
        qs = self.filter_queryset(self.get_queryset()).order_by(
            '-year',
            'iso3',
        )

        wb = Workbook(write_only=True)
        # Tab 1
        ws = wb.create_sheet('1_Displacement_data')

        request_cause = request.GET.get('cause')

        if request_cause and request_cause.lower() == 'conflict':
            self.export_conflicts(ws, qs)
        if request_cause and request_cause.lower() == 'disaster':
            self.export_disasters(ws, qs)
        else:
            self.export_displacements(ws, qs)
        # Tab 2
        ws2 = wb.create_sheet('2_Context_Displacement_data')
        ws2.append([
            'ISO3',
            'Year',
            'Figure cause',
            'Figure category',
            'Description',
            'Figures',
            'Figures rounded',
        ])
        pfa_qs = PublicFigureAnalysisFilterSet(
            data=self.request.query_params,
            queryset=PublicFigureAnalysis.objects.all()
        ).qs.order_by('iso3', 'year')
        for item in pfa_qs:
            ws2.append([
                item.iso3,
                item.year,
                item.figure_cause.label,
                item.figure_category.label,
                item.description,
                item.figures,
                item.figures_rounded,
            ])
        # Tab 3
        ws3 = wb.create_sheet('3_IDPs_SADD_estimates')
        ws3.append([
            'ISO3',
            'Country',
            'Year',
            'Sex',
            'Cause',
            '0-4',
            '5-11',
            '12-17',
            '18-59',
            '60+',
        ])
        idps_sadd_qs = IdpsSaddEstimateFilter(
            data=self.request.query_params,
            queryset=IdpsSaddEstimate.objects.all(),
        ).qs.order_by('iso3', 'year')
        for item in idps_sadd_qs:
            ws3.append([
                item.iso3,
                item.country_name,
                item.year,
                item.sex,
                item.cause.label,
                item.zero_to_four,
                item.five_to_eleven,
                item.twelve_to_seventeen,
                item.eighteen_to_fiftynine,
                item.sixty_plus,
            ])
        # Tab 4
        ws4 = wb.create_sheet('README')
        readme_text = [
            ['TITLE: Global Internal Displacement Database (GIDD)'],
            [],
            ['FILENAME: IDMC_Internal_Displacement_Conflict-Violence_Disasters'],
            [],
            ['SOURCE: Internal Displacement Monitoring Centre (IDMC)'],
            [],

            [f'DATE EXTRACTED: {datetime.now().strftime("%B %d, %Y")}'],
            [],
            [f'LAST UPDATE: {StatusLog.last_release_date()}'],
            [],
            ['DESCRIPTION:'],
            [
                'The Internal Displacement Monitoring Centre (IDMC) monitors internal displacement events globally, '
                'triggered by disasters, conflict, and other forms of violence. It gathers and analyses both structured '
                'and unstructured secondary data from diverse sources—including government agencies, UN agencies, the '
                'International Federation of the Red Cross and Red Crescent, and the media.\n'
                '\n'
                'IDMC analysts rigorously analyse and triangulate all reported data. The data undergo thorough quality '
                'control processes, involving engagement with primary data collectors for peer review and validation. '
                'This meticulous approach guarantees that the data reported by IDMC reflects high accuracy.\n'
                '\n'
                'The data in the Global Internal Displacement Database (GIDD) is annually validated and peer-reviewed, '
                'having passed through various quality control processes in consultation with different UN agencies, '
                'goverments and local data providers.\n'
                '\n'
                'The GIDD database documents displacement due to conflict from 2009 to 2023 and disaster-induced '
                'displacement from 2008 to 2023. For detailed definitions and more comprehensive descriptions, please '
                'refer to the IDMC Monitoring Tools (https://www.internal-displacement.org/monitoring-tools).\n'
            ],
            [],
            ['KEY DEFINITIONS:'],
            [],
            [
                'Internal Displacements (flows): This metric represents the number of internal displacements, or '
                'internal displacement population flows, reported from January 1st to December 31st of a reporting year.'
                'This figure may include individuals who are displaced multiple times during the year by different events.'
            ],
            [
                'Total number of Internally Displaced Persons (IDPs) (stocks): This metric represents the total number '
                'of people living in situations of internal displacement as of the end of the reporting year, '
                'specifically on December 31st of each year.'
            ],
            [
                'Conflict displacement: Refers to situations where people are forced to leave their homes or places of '
                'habitual residence as a result or in order to avoid the impact of armed conflict, communal violence '
                'and criminal violence.'
            ],
            [
                'Disaster displacement: Refers to situations where people are forced to leave their homes or places of '
                'habitual residence as a result, or in anticipation of the negative impact of natural hazards.'
            ],
            [
                'Disaster: A serious disruption of the functioning of a community or a society involving widespread '
                'human, material, economic or environmental losses and impacts, which exceeds the ability of the '
                'affected community or society to cope using its own resources (UNSDR).'
            ],
            [],
            [
                'USE LICENSE: This content is licensed under CC BY-NC. Detailed licensing information is available at '
                'Creative Commons License (See: https://creativecommons.org/licenses/by-nc/4.0/).'
            ],
            [],
            [
                'COVERAGE: Global. The GIDD provides data on internal displacement caused by conflict since 2009. This '
                'includes information on both metrics: internal displacements and the total number of IDPs. Data on '
                'internal displacements triggered by disasters dates back to 2008, and the metrics on the total number '
                'of IDPs from disaster-related events are available from 2019 onwards.'
            ],
            [],
            ['CITATION:'],
            [
                'All derived work from IDMC data could cite IDMC following this example: Internal Displacement '
                'Monitoring Centre. Global Internal Displacement Database. IDMC (2023). Available at: '
                'https://www.internal-displacement.org/database/displacement-data/ (Accessed: [date of access]).'
            ],
            [],
            ['CONTACT: info@idmc.ch'],
            [],
        ]

        for item in readme_text:
            ws4.append(item)

        ws4.append(['DATA DESCRIPTION: 1_Displacement_data table'])
        readme_text_2 = [
            [
                'This dataset offers annually validated data on internal displacement caused by disasters, conflicts, '
                'and other situations of violence, as compiled and reported by the Internal Displacement Monitoring '
                'Centre (IDMC).'
            ],
            [],
            ['ISO3: Represents the ISO 3166-1 alpha-3 code. The code \'AB9\' is assigned to the Abyei Area.'],
            ['Country / Territory: Short name of the country or territory.'],
            ['Year: Indicates the year for which displacement data are reported.'],
            [
                'Conflict Total number of IDPs: Total number of IDPs (rounded figures at '
                'the national level), as a result of conflict and violence as of the end of '
                'the reporting year. Units are recorded as \'People\'.'
            ],
            [
                'Conflict Total number of IDPs raw: Total number of IDPs (not rounded), as a result of conflict and '
                'violence as of the end of the reporting year. Units are recorded as \'People\'.'
            ],
            [
                'Conflict Internal Displacements: Total number of internal displacements reported (rounded figures at '
                'national level), as a result of conflict and violence over the reporting year. Units are recorded as '
                '\'internal displacement flows\' or \'internal displacement movements.\' '
            ],
            [
                'Conflict Internal Displacements raw: Total number of internal displacements '
                'reported (not rounded), as a result of conflict and violence over the reporting year.  Units are '
                'recorded as \'internal displacement flows\' or \'internal displacement movements.\' '
            ],
            [
                'Disaster Internal Displacements: Total number of internal displacements reported '
                '(rounded figures at national level), as a result of disasters over the reporting year. Units are '
                'recorded as \'internal displacement flows\' or \'internal displacement movements.\' '
            ],
            [
                'Disaster Internal Displacements raw: Total number of internal displacements reported (not rounded), '
                'as a result of disasters over the reporting year. Units are recorded as \'internal displacement flows\''
                ' or \'internal displacement movements.\' '
            ],
            [
                'Disaster Total number of IDPs: Total number of IDPs (rounded figures at '
                'national level), as a result, of disasters as of the end of the reporting year. '
                'Units are recorded as \'People\'.'
            ],
            [
                'Disaster Total number of IDPs raw: Total number of IDPs (not rounded), as a result, of disasters as of '
                'the end of the reporting year. Units are recorded as \'People\'.'
            ],
        ]
        ws4.append([])
        for item in readme_text_2:
            ws4.append(item)
        ws4.append([])
        ws4.append([
            'DATA DESCRIPTION: 2_Context_Displacement_data table'
        ])
        readme_text_3 = [
            [
                'This dataset provides contextual information and analysis documented by IDMC analysts. '
                'It captures flags related to methodology, caveats, sources, and challenges identified for each '
                'metric, reporting year, and country.'
            ],
            [],
            ['ISO3: Represents the ISO 3166-1 alpha-3 code, the code \'AB9\' is assigned to the Abyei Area.'],
            ['Year: Indicates the year for which displacement data are reported.'],
            ['Figure_Cause_Name: Identifies the trigger of displacement, such as conflict or disasters.'],
            [
                'Figure_Category_Name: Categorizes the type of displacement metric. It details values for '
                'Internal Displacements (internal displacement flows) and Total Number of IDPs (internal displacement '
                'stocks), as defined earlier in this document.'
            ],
            [
                'Description: Provides contextual information about the data, including sources and data limitations. '
                'It is essential for representing the analysis conducted by IDMC analysts. This field also details the '
                'methodology used, descriptions of sources, and outlines any caveats and challenges identified with '
                'the displacement figures reported.'
            ],
            [
                'Figures: Represents the total number of internal displacements or IDPs. For internal displacements, '
                'units are recorded as \'internal displacement flows\' or \'internal displacement movements.\' For total '
                'number of IDPs, units reflect the total number of people living in displacement.'
            ],
            [
                'Figures_Rounded: Displays rounded figures to provide a simplified view of the data that matches the '
                'figures reported in the Global Report on Internal Displacement (GRID).'
            ],
        ]
        ws4.append([])
        for item in readme_text_3:
            ws4.append(item)
        readme_text_4 = [
            ['ISO3: Represents the ISO 3166-1 alpha-3 code, the code \'AB9\' is assigned to the Abyei Area.'],
            ['Country: Short name of the country or territory.'],
            ['Geographical region: Corresponds to IDMC\'s geographical regions'],
            ['Year: The year for which displacement figures are reported.'],
            [
                'Sex : This field contains information on Female, Male and Both Sexes categories following the United '
                'Nations Department of Economic and Social Affairs (UN DESA) classifications. '
            ],
            ['Cause:  Identifies the trigger of displacement, such as conflict or disasters.'],
            ['Age_0_4: Represents the age cohort from newborns to 4 years old.'],
            ['Age_5_11: Represents children aged 5 to 11 years.'],
            ['Age_12_17: Represents adolescents aged 12 to 17 years.'],
            ['Age_18_59: Represents adults aged 18 to 59 years.'],
            ['Age_60_plus: Represents the population aged 60 years and older.'],
        ]
        ws4.append([])
        ws4.append([
            'DATA DESCRIPTION: 3_IDPs_SADD_estimates table'
        ])
        ws4.append([])
        ws4.append([
            'Sex and Age Disaggregated Data (SADD) for displacement associated with conflict or disasters is often '
            'scarce. One way to estimate it is to use SADD available at the national level. IDMC employs United '
            'Nations Population Estimates and Projections to break down the number of internally displaced people by '
            'sex and age. The methodology and limitations of this approach are described on IDMC’s website at: '
            'https://www.internal-displacement.org/monitoring-tools'
        ])
        ws4.append([])
        for item in readme_text_4:
            ws4.append(item)

        response = HttpResponse(content=save_virtual_workbook(wb))
        filename = 'IDMC_Internal_Displacement_Conflict-Violence_Disasters.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        response['Content-Type'] = 'application/octet-stream'
        return response


class DisaggregationViewSet(ListOnlyViewSetMixin):
    queryset = GiddFigure.objects.all()
    serializer_class = DisaggregationSerializer
    filter_backends = (DjangoFilterBackend, )
    filterset_class = DisaggregationFilterSet
    pagination_class = None

    def _get_category(self, category) -> typing.Optional[str]:
        if category is None:
            return
        return Figure.FIGURE_CATEGORY_TYPES.get(category).label

    def _get_cause(self, cause) -> typing.Optional[str]:
        if cause is None:
            return
        return Crisis.CRISIS_TYPE.get(cause).label

    def _get_date_accuracy(self, accuracy) -> typing.Optional[str]:
        if accuracy is None:
            return
        return DATE_ACCURACY.get(accuracy).label

    def _get_displacement_occurred(self, displacement_occurred) -> str:
        if displacement_occurred is not None:
            return DisasterViewSet.get_displacement_status([displacement_occurred])
        return ""

    def _get_unit(self, unit) -> typing.Optional[str]:
        if unit is None:
            return None
        return Figure.UNIT.get(unit).label

    def extract_event_data(
        self,
        event_code: typing.List[typing.Tuple[str]],
        event_code_type: typing.List[typing.Tuple[int]],
        event_code_iso3: typing.List[typing.Tuple[str]],
        filter_iso3: str
    ) -> str:
        event_code_components = [
            event_code,
            event_code_type,
            event_code_iso3
        ]
        transposed_components = zip(*event_code_components)

        return EXTERNAL_ARRAY_SEPARATOR.join(
            EXTERNAL_FIELD_SEPARATOR.join([loc[0], _get_event_code_label(loc[1])])
            for loc in transposed_components
            if loc[2] == filter_iso3
        )

    def extract_event_data_raw(
        self,
        event_code: typing.List[typing.Tuple[str]],
        event_code_type: typing.List[typing.Tuple[int]],
        event_code_iso3: typing.List[typing.Tuple[str]],
        filter_iso3: str
    ) -> str:
        event_code_components = [
            event_code,
            event_code_type,
            event_code_iso3
        ]
        transposed_components = zip(*event_code_components)

        return [
            [loc[0], _get_event_code_label(loc[1])]
            for loc in transposed_components
            if loc[2] == filter_iso3
        ]

    def _export_disaggregated_geojson(self, qs):
        def format_coordinate(coordinate: str) -> typing.Tuple[float, float]:
            lat, lng = coordinate.split(', ')
            return (float(lng), float(lat))

        def format_coordinates(coordinates: typing.List[str]):
            return [format_coordinate(x) for x in coordinates]

        qs = qs.exclude(
            Q(locations_coordinates__isnull=True) |
            Q(locations_coordinates=[])
        ).annotate(
            event_main_trigger=Case(
                When(
                    gidd_event__cause=Crisis.CRISIS_TYPE.CONFLICT,
                    then=F('gidd_event__violence__name')
                ),
                When(
                    gidd_event__cause=Crisis.CRISIS_TYPE.DISASTER,
                    then=F('gidd_event__disaster_sub_type__name')
                ),
                When(
                    gidd_event__cause=Crisis.CRISIS_TYPE.OTHER,
                    then=F('gidd_event__other_sub_type__name')
                ),
                output_field=models.CharField()
            ),
        )
        now = timezone.now().strftime("%B %d, %Y")

        # Determine the filename based on filters
        filter_cause = self.request.query_params.get('cause', '').lower()
        filename_map = {
            Crisis.CRISIS_TYPE.CONFLICT.name.lower(): 'IDMC_GIDD_Conflict_Internal_Displacement_Disaggregated',
            Crisis.CRISIS_TYPE.DISASTER.name.lower(): 'IDMC_GIDD_Disasters_Internal_Displacement_Disaggregated'
        }
        filename = filename_map.get(filter_cause, 'IDMC_GIDD_Internal_Displacement_Disaggregated')

        readme_text = (
            "TITLE: Disasters Global Internal Displacement Database (GIDD)\n"
            "\n"
            f"FILENAME: {filename}\n"
            "\n"
            "SOURCE: Internal Displacement Monitoring Centre (IDMC)\n"
            "\n"
            f"DATE EXTRACTED: {now}\n"
            "\n"
            f"LAST UPDATE: {StatusLog.last_release_date()}\n"
            "\n"
            "DESCRIPTION:\n"
            "The Internal Displacement Monitoring Centre (IDMC) monitors internal displacement events globally, "
            "triggered by disasters, conflict, and other forms of violence. It gathers and analyses both structured and "
            "unstructured secondary data from diverse sources—including government agencies, UN agencies, "
            "the International Federation of the Red Cross and Red Crescent, and the media.\n"
            "\n"
            "IDMC analysts rigorously analyse and triangulate all reported data. The data undergo thorough "
            "quality control processes, involving engagement with primary data collectors for peer review and validation. "
            "This meticulous approach guarantees that the data reported by IDMC reflects high accuracy.\n"
            "\n"
            "The data in the Global Internal Displacement Database (GIDD) is annually validated and peer-reviewed, "
            "having passed through various quality control processes in consultation with different UN agencies, "
            "goverments and local data providers.\n"
            "\n"
            "The GIDD database documents displacement due to conflict from 2009 to 2023 and disaster-induced "
            "displacement from 2008 to 2023. For detailed definitions and more comprehensive descriptions, "
            "please refer to the IDMC Monitoring Tools (https://www.internal-displacement.org/monitoring-tools).\n"
            "\n"
            "KEY DEFINITIONS:\n"
            "\n"
            "Internal Displacements (flows): This metric represents the number of internal displacements, "
            "or internal displacement population flows, reported from January 1st to December 31st of a reporting year. "
            "This figure may include individuals who are displaced multiple times during the year by different events.\n"
            "Total number of Internally Displaced Persons (IDPs) (stocks): This metric represents the total number "
            "of people living in situations of internal displacement as of the end of the reporting year, "
            "specifically on December 31st of each year.\n"
            "Conflict displacement: Refers to situations where people are forced to leave their homes or places "
            " of habitual residence as a result or in order to avoid the impact of armed conflict, communal violence\n"
            "and criminal violence.\n"
            "Disaster displacement: Refers to situations where people are forced to leave their homes or places "
            "of habitual residence as a result, or in anticipation of the negative impact of natural hazards.\n"
            "Disaster: A serious disruption of the functioning of a community or a society involving widespread "
            "human, material, economic or environmental losses and impacts, which exceeds the ability of the\n"
            "affected community or society to cope using its own resources (UNSDR).\n"
            "\n"
            "USE LICENSE: This content is licensed under CC BY-NC. Detailed licensing information is available at "
            "Creative Commons License (See: https://creativecommons.org/licenses/by-nc/4.0/).\n"
            "\n"
            "COVERAGE: Global. The GIDD provides data on internal displacement caused by conflict since 2009. This "
            "includes information on both metrics: internal displacements and the total number of IDPs. Data on internal "
            "displacements triggered by disasters dates back to 2008, and the metrics on the total number of IDPs from "
            "disaster-related events are available from 2019 onwards.\n"
            "\n"
            "CITATION:\n"
            "All derived work from IDMC data could cite IDMC following this example: Internal Displacement Monitoring "
            "Centre. Global Internal Displacement Database. IDMC (2023). "
            "Available at: https://www.internal-displacement.org/database/displacement-data/ (Accessed: [date of access]).\n"
            "\n"
            "CONTACT: info@idmc.ch\n"
            "\n"
            "DATA DESCRIPTION: 1_Disaggregated_Data table\n"
            "\n"
            "ID: IDMC figure unique identifier.\n"
            "ISO3: Represents the ISO 3166-1 alpha-3 code. The code 'AB9' is assigned to the Abyei Area. "
            "uncertanty or accuracy of start date\n"
            "Country / Territory: Short name of the country or territory.\n"
            "Geographical region: Corresponds to IDMC's geographical regions.\n"
            "Figure cause:  Identifies the trigger of displacement, such as conflict or disasters.\n"
            "Year: Indicates the year for which displacement data are reported.\n"
            "Figure category:  Categorizes the type of displacement metric. It details values for Internal Displacements "
            "(internal displacement flows) and Total Number of IDPs (internal displacement stocks), as "
            "defined earlier in this document.\n"
            "Total figures: Represents the total number of internal displacements or IDPs. For internal displacements, "
            "units are recorded as 'internal displacement flows' or 'internal displacement movements.' "
            "For total number of IDPs, units reflect the total number of people living in displacement.\n"
            "Reported figures: This field represents the values reported by the original source. Figures can be reported "
            "either in terms of households or individual counts.\n"
            "Figure unit: This field specifies the type of unit reported in the 'Reported' column. Possible values include "
            "'households' or 'people'. The catogy people include  'internal displacement flows' or 'internal "
            "displacement movements.'\n"
            "Household size: This metric represents the average number of individuals per household. It is calculated using "
            "data from various sources, including the United Nations Department of Economic and Social Affairs "
            "(UNDESA), national statistical offices, and estimates from local primary data providers shared with IDMC.\n"
            "Hazard Category: Hazard category based on the CRED EM-DAT classification.\n"
            "Hazard sub category: Hazard sub category based on the CRED EM-DAT classification.\n"
            "Hazard Type: Hazard type as categorized by CRED EM-DAT.\n"
            "Hazard Sub-Type: Specific sub-type of the hazard based on CRED EM-DAT.\n"
            "Start date: Start date of displacement flow.\n"
            "Start date accuracy: Uncertanty or accuracy of start date.\n"
            "End date: End date of thedisplacement flow.\n"
            "End date accuracy: Uncertanty or accuracy of end date.\n"
            "Stock date: This field indicates the year in which the data for the IDP metric (total number of internally "
            "displaced persons or stocks) was collected.\n"
            "Stock date accuracy: Uncertanty or accuracy of stock date.\n"
            "Stock reporting date: This field reflects the year IDMC uses to report the total number of internally "
            "displaced persons (IDPs). It represents the IDMC reporting year, which may not coincide with the actual "
            "data collection year. Given the protracted nature of displacement, annual updates on the total number of "
            "IDPs may not always be available. To maintain accuracy in reporting, IDMC relies on the most recent verified "
            "data until evidence shows that the displaced population has achieved a durable solution.\n"
            "Publishers: Organizations responsible for distributing and disseminating internal displacement data\n"
            "Sources:  This field lists the names of the primary data providers or the original sources for the internal "
            "displacement data reported by IDMC.\n"
            "Sources type: This field categorizes the type of source as defined by IDMC.\n"
            "Event ID: Unique identifier for events as assigned by IDMC.\n"
            "Event name: This field includes the event's coded name, which is based on the country, type of hazard, "
            "location, and start date. It also incorporates the common or official name of the event, when available.\n"
            "Event cause: Identifies the trigger of displacement, such as conflict or disasters.\n"
            "Event main trigger: This field identifies the primary hazard subtype or conflict type that initiated "
            "the event, serving as the main driver of a disaster or conflict. For disasters, associated fields such "
            "as \"Hazard Category,\" \"Hazard Subcategory,\" \"Hazard Type,\" and \"Hazard Sub-Type\" detail the cascading "
            "impacts stemming from this main trigger. For instance, a tropical storm identified as the main driver of "
            "displacement might lead to reports in \"Hazard Sub-Type\" of floods, landslides, and other related disaster "
            "types arising from the initial hazard.\n"
            "Event start date: Event or hazard start date.\n"
            "Event end date: Event or hazard end date date.\n"
            "Event start date accuracy: Uncertanty or accuracy of event start date.\n"
            "Event end date accuracy: Uncertanty or accuracy of event end date.\n"
            "Is housing destruction: This field  indicates whether the displacement data includes individuals displaced "
            "by housing destruction. Values are \"Yes\" if the data reflects households whose homes were destroyed and "
            "\"No\" otherwise. This field relies on the data specified in \"Reported Figures\" "
            "and is linked to the \"Unit\" "
            "of measurement used, which in this context refers to houses destroyed.\n"
            "Violence type: This field categorizes the type of violence using IDMC's typology, which aligns with "
            "international classifications. The categories include - International Armed Conflict (IAC): Refers to "
            "armed conflict between two or more states. - Non-International Armed Conflict (NIAC): Refers to protracted "
            "armed conflict occurring within the territory of a single state between its government and non-state armed "
            "groups, or between such groups themselves. - Unclear/Unknown: Indicates situations where the type of violence "
            "is not definitively categorized due to limited information.\n"
            "Event codes (Code:Type): Unique codes such as the GLIDE number and other database-specific codes used "
            "to identify and track specific events across various databases.\n"
            "Locations name: This field indicates the names of locations where displacement incidents have been "
            "reported. It's important to note that this field may exhibit a many-to-one relationship, signifying that "
            "multiple location names could be associated with a single reported figure, preventing disaggregation by "
            "individual location. This becomes particularly relevant in geospatial analysis, where Geographic Information "
            "System (GIS) software may interpret these multi-point entities as single data points, potentially leading to "
            "the inadvertent double-counting of figures. To mitigate this issue, it's advisable to preprocess the dataset "
            "by either dividing the total figure by the number of locations or "
            "distributing the \"Total figures\" values based "
            "on a weighting factor such as population density. This ensures a more accurate representation "
            "of the displacement data across individual locations and prevents duplication of figures during analysis.\n"
            "Locations coordinates: This field contains geographic coordinates representing the reported locations. "
            "Please note that this field contains multipoints  meaning that multiple locations may represent one figures. "
            "It's important to note that this field may exhibit a many-to-one relationship, signifying that multiple "
            "location names could be associated with a single reported figure, preventing disaggregation by individual "
            "location. This becomes particularly relevant in geospatial analysis, where Geographic Information System (GIS) "
            "software may interpret these multi-point entities as single data points, potentially leading to the "
            "inadvertent double-counting of figures. To mitigate this issue, it's advisable to preprocess the dataset "
            "by either dividing the total figure by the number of locations or distributing the \"Total figures\" "
            "values based on a weighting factor such as population density. This ensures a more accurate representation "
            "of the displacement data across "
            "individual locations and prevents duplication of figures during analysis.\n"
            "Locations accuracy:  This field indicates the estimated precision of the reported locations. It serves "
            "as a clue to the likely administrative unit level (e.g., country, state, district) used for reporting.\n"
            "Locations type: This field specifies the type of displacement location within a reported event. It can "
            "indicate,  Origin: The place where people were displaced from. Destination: The location where displaced "
            "people arrived. Both: In some cases, both origin and destination information might be included.It's crucial "
            "to note that different locations reported for a single figure may pertain to both the origin and destination "
            "of displacement incidents. This distinction is particularly salient in geospatial analysis, where Geographic "
            "Information System (GIS) software may interpret these "
            "multi-point entities as singular data points, potentially "
            "resulting in inadvertent double-counting of figures. To mitigate this issue, it is recommended to preprocess "
            "the dataset prior to GIS analysis to ensure accurate representation and avoid duplication of figures.\n"
            "Displacement occurred: Displacement Occurred: This field contains values that represent if preventive "
            "evacuations were reported. These evacuations are the result of existing early warning systems.\n"
        )
        feature_collection = {
            "type": "FeatureCollection",
            "readme": readme_text,
            "features": []
        }

        for item in qs:
            feature = {
                "type": "Feature",
                "geometry": {
                    "type": "MultiPoint",
                    "coordinates": format_coordinates(item.locations_coordinates),
                },
                "properties": remove_null_from_dict({
                    "ID": item.figure_id,
                    "ISO3": item.iso3,
                    "Country": item.country_name,
                    "Geographical region": item.geographical_region_name,
                    "Figure cause": self._get_cause(item.cause),
                    "Year": item.year,
                    "Figure category": self._get_category(item.category),
                    "Figure unit": self._get_unit(item.unit),
                    "Reported figures": item.reported,
                    "Household size": item.household_size,
                    "Total figures": item.total_figures,
                    "Hazard category": item.disaster_category_name,
                    "Hazard sub category": item.disaster_sub_category_name,
                    "Hazard type": item.disaster_type_name,
                    "Hazard sub type": item.disaster_sub_type_name,
                    "Violence type": item.violence_name,
                    "Other event sub type": item.other_sub_type_name,
                    "Start date": item.start_date,
                    "Start date accuracy": self._get_date_accuracy(item.start_date_accuracy),
                    "End date": item.end_date,
                    "End date accuracy": self._get_date_accuracy(item.end_date_accuracy),
                    "Stock date": item.stock_date,
                    "Stock date accuracy": self._get_date_accuracy(item.stock_date_accuracy),
                    "Stock reporting date": item.stock_reporting_date,
                    "Publishers": item.publishers,
                    "Sources": item.sources,
                    "Sources type": item.sources_type,
                    "Event ID": item.gidd_event.event_id,
                    "Event name": item.gidd_event.name,
                    "Event cause": self._get_cause(item.gidd_event.cause),
                    "Event main trigger": item.event_main_trigger,
                    "Event start date": item.gidd_event.start_date,
                    "Event end date": item.gidd_event.end_date,
                    "Event start date accuracy": self._get_date_accuracy(item.gidd_event.start_date_accuracy),
                    "Event end date accuracy": self._get_date_accuracy(item.gidd_event.end_date_accuracy),
                    "Is housing destruction": "Yes" if item.is_housing_destruction is not None else "No",
                    "Event codes (Code:Type)": self.extract_event_data_raw(
                        item.gidd_event.event_codes,
                        item.gidd_event.event_codes_type,
                        item.gidd_event.event_codes_iso3,
                        item.iso3,
                    ),
                    "Locations name": item.locations_names,
                    "Locations accuracy": [_get_location_accuracy_label(x) for x in item.locations_accuracy],
                    "Locations type": [_get_location_type_label(x) for x in item.locations_type],
                    "Displacement occurred": self._get_displacement_occurred(item.displacement_occurred),
                })
            }
            feature_collection['features'].append(feature)

        feature_collection = json.dumps(feature_collection, cls=DjangoJSONEncoder)
        response = HttpResponse(content=feature_collection, content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename={filename}.geojson'
        return response

    def _export_disaggregated_excel(self, qs):
        wb = Workbook(write_only=True)

        # Determine the filename based on filters
        filter_cause = self.request.query_params.get('cause', '').lower()
        filename_map = {
            Crisis.CRISIS_TYPE.CONFLICT.name.lower(): 'IDMC_GIDD_Conflict_Internal_Displacement_Disaggregated',
            Crisis.CRISIS_TYPE.DISASTER.name.lower(): 'IDMC_GIDD_Disasters_Internal_Displacement_Disaggregated'
        }
        filename = filename_map.get(filter_cause, 'IDMC_GIDD_Internal_Displacement_Disaggregated')

        ws = wb.create_sheet('1_Disaggregated_Data')
        ws.append([
            'ID',
            'ISO3',
            'Country',
            'Geographical region',
            'Figure cause',
            'Year',
            'Figure category',
            'Figure unit',
            'Reported figures',
            'Household size',
            'Total figures',
            'Hazard category',
            'Hazard sub category',
            'Hazard type',
            'Hazard sub type',
            'Violence type',
            'Other event sub type',
            'Start date',
            'Start date accuracy',
            'End date',
            'End date accuracy',
            'Stock date',
            'Stock date accuracy',
            'Stock reporting date',
            'Publishers',
            'Sources',
            'Sources type',
            'Event ID',
            'Event name',
            'Event cause',
            'Event main trigger',
            'Event start date',
            'Event end date',
            'Event start date accuracy',
            'Event end date accuracy',
            'Is housing destruction',
            'Event codes (Code:Type)',
            'Locations coordinates',
            'Locations name',
            'Locations accuracy',
            'Locations type',
            'Displacement occurred',
        ])

        # Tab 2
        ws2 = wb.create_sheet('2_Context_Displacement_data')
        ws2.append([
            'ISO3',
            'Year',
            'Figure cause',
            'Figure category',
            'Description',
            'Figures',
            'Figures rounded',
        ])

        pfa_qs = DisaggregationPublicFigureAnalysisFilterSet(
            data=self.request.query_params,
            queryset=PublicFigureAnalysis.objects.filter(year__gte=2023)
        ).qs.order_by('iso3', 'year', 'id')

        for item in pfa_qs:
            ws2.append([
                item.iso3,
                item.year,
                item.figure_cause.label,
                item.figure_category.label,
                item.description,
                item.figures,
                item.figures_rounded,
            ])

        # README TAB
        ws3 = wb.create_sheet('README')
        readme_text = [
            ['TITLE: Disasters Global Internal Displacement Database (GIDD)'],
            [],
            [f'FILENAME: {filename}'],
            [],
            ['SOURCE: Internal Displacement Monitoring Centre (IDMC)'],
            [],
            [f'DATE EXTRACTED: {timezone.now().strftime("%B %d, %Y")}'],
            [],
            [f'LAST UPDATE: {StatusLog.last_release_date()}'],
            [],
            ['DESCRIPTION:'],
            [
                'The Internal Displacement Monitoring Centre (IDMC) monitors internal displacement events globally, '
                'triggered by disasters, conflict, and other forms of violence. It gathers and analyses both '
                'structured and unstructured secondary data from diverse sources—including government agencies, '
                'UN agencies, the International Federation of the Red Cross and Red Crescent, and the media.'
            ],
            [],
            [
                'IDMC analysts rigorously analyse and triangulate all reported data. The data undergo thorough quality '
                'control processes, involving engagement with primary data collectors for peer review and validation. '
                'This meticulous approach guarantees that the data reported by IDMC reflects high accuracy.'
            ],
            [],
            [
                'The data in the Global Internal Displacement Database (GIDD) is annually validated and peer-reviewed, '
                'having passed through various quality control processes in consultation with different UN agencies, '
                'goverments and local data providers.'
            ],
            [],
            [
                'The GIDD database documents displacement due to conflict from 2009 to 2023 and disaster-induced '
                'displacement from 2008 to 2023. For detailed definitions and more comprehensive descriptions, please '
                'refer to the IDMC Monitoring Tools (https://www.internal-displacement.org/monitoring-tools).'
            ],
            [],
            ['KEY DEFINITIONS:'],
            [],
            [
                'Internal Displacements (flows): This metric represents the number of internal displacements, or '
                'internal displacement population flows, reported from January 1st to December 31st of a reporting year. '
                'This figure may include individuals who are displaced multiple times during the year by different events.'
            ],
            [
                'Total number of Internally Displaced Persons (IDPs) (stocks): This metric represents the total number '
                'of people living in situations of internal displacement as of the end of the reporting year, '
                'specifically on December 31st of each year.'
            ],
            [
                'Conflict displacement: Refers to situations where people are forced to leave their homes or places of '
                'habitual residence as a result or in order to avoid the impact of armed conflict, communal violence '
                'and criminal violence.'
            ],
            [
                'Disaster displacement: Refers to situations where people are forced to leave their homes or places of '
                'habitual residence as a result, or in anticipation of the negative impact of natural hazards.'
            ],
            [
                'Disaster: A serious disruption of the functioning of a community or a society involving widespread '
                'human, material, economic or environmental losses and impacts, which exceeds the ability of the '
                'affected community or society to cope using its own resources (UNSDR).'
            ],
            [],
            [
                'USE LICENSE: This content is licensed under CC BY-NC. Detailed licensing information is available at '
                'Creative Commons License (See: https://creativecommons.org/licenses/by-nc/4.0/).'
            ],
            [],
            [
                'COVERAGE: Global. The GIDD provides data on internal displacement caused by conflict since 2009. This '
                'includes information on both metrics: internal displacements and the total number of IDPs. Data on '
                'internal displacements triggered by disasters dates back to 2008, and the metrics on the total number '
                'of IDPs from disaster-related events are available from 2019 onwards.'
            ],
            [],
            ['CITATION: '],
            [
                'All derived work from IDMC data could cite IDMC following this example: Internal Displacement '
                'Monitoring Centre. Global Internal Displacement Database. IDMC (2023). Available at: '
                'https://www.internal-displacement.org/database/displacement-data/ (Accessed: [date of access]).'
            ],
            [],
            ['CONTACT: info@idmc.ch'],
        ]

        for item in readme_text:
            ws3.append(item)
        ws3.append([])
        ws3.append(['DATA DESCRIPTION: 1_Disaggregated_Data table'])
        ws3.append([])

        data_description_1 = [
            ["ID: IDMC figure unique identifier."],
            ["ISO3: Represents the ISO 3166-1 alpha-3 code. The code 'AB9' is assigned to the Abyei Area."],
            ["Country / Territory: Short name of the country or territory."],
            ["Geographical region: Corresponds to IDMC's geographical regions."],
            ["Figure cause:  Identifies the trigger of displacement, such as conflict or disasters."],
            ["Year: Indicates the year for which displacement data are reported."],
            [
                "Figure category:  Categorizes the type of displacement metric. It details values for Internal "
                "Displacements (internal displacement flows) and Total Number of IDPs (internal displacement stocks), "
                "as defined earlier in this document."
            ],
            [
                "Total figures: Represents the total number of internal displacements or IDPs. For internal "
                "displacements, units are recorded as 'internal displacement flows' or 'internal displacement "
                "movements.' For total number of IDPs, units reflect the total number of people living in displacement."
            ],
            [
                "Reported figures: This field represents the values reported by the original source. Figures can be "
                "reported either in terms of households or individual counts."
            ],
            [
                "Figure unit: This field specifies the type of unit reported in the 'Reported' column. Possible values "
                "include 'households' or 'people'. The catogy people include  'internal displacement flows' or 'internal"
                " displacement movements.'"
            ],
            [
                "Household size: This metric represents the average number of individuals per household. It is "
                "calculated using data from various sources, including the United Nations Department of Economic and "
                "Social Affairs (UNDESA), national statistical offices, and estimates from local primary data providers "
                "shared with IDMC."
            ],
            ["Hazard Category: Hazard category based on the CRED EM-DAT classification."],
            ["Hazard sub category: Hazard sub category based on the CRED EM-DAT classification."],
            ["Hazard Type: Hazard type as categorized by CRED EM-DAT."],
            ["Hazard Sub-Type: Specific sub-type of the hazard based on CRED EM-DAT."],
            ["Start date: Start date of displacement flow."],
            ["Start date accuracy: Uncertainty or accuracy of start date."],
            ["End date: End date of thedisplacement flow."],
            ["End date accuracy: Uncertainty or accuracy of end date."],
            [
                "Stock date: This field indicates the year in which the data for the IDP metric (total number of "
                "internally displaced persons or stocks) was collected."
            ],
            ["Stock date accuracy: Uncertainty or accuracy of stock date."],
            [
                "Stock reporting date: This field reflects the year IDMC uses to report the total number of internally "
                "displaced persons (IDPs). It represents the IDMC reporting year, which may not coincide with the "
                "actual data collection year. Given the protracted nature of displacement, annual updates on the total "
                "number of IDPs may not always be available. To maintain accuracy in reporting, IDMC relies on the "
                "most recent verified data until evidence shows that the displaced population has achieved a durable "
                "solution."
            ],
            ["Publishers: Organizations responsible for distributing and disseminating internal displacement data"],
            [
                "Sources: This field lists the names of the primary data providers or the original sources for the "
                "internal displacement data reported by IDMC."
            ],
            ["Sources type: This field categorizes the type of source as defined by IDMC."],
            ["Event ID: Unique identifier for events as assigned by IDMC."],
            [
                "Event name: This field includes the event's coded name, which is based on the country, type of hazard, "
                "location, and start date. It also incorporates the common or official name of the event, when available."
            ],
            ["Event cause: Identifies the trigger of displacement, such as conflict or disasters."],
            [
                "Event main trigger: This field identifies the primary hazard subtype or conflict type that initiated "
                "the event, serving as the main driver of a disaster or conflict. For disasters, associated fields such "
                "as \"Hazard Category\", \"Hazard Subcategory\", \"Hazard Type\", and \"Hazard Sub-Type\" detail the "
                "cascading impacts stemming from this main trigger. For instance, a tropical storm identified as the "
                "main driver of displacement might lead to reports in \"Hazard Sub-Type\" of floods, landslides, and "
                "other related disaster types arising from the initial hazard."
            ],
            ["Event start date: Event or hazard start date."],
            ["Event end date: Event or hazard end date date."],
            ["Event start date accuracy: Uncertainty or accuracy of event start date."],
            ["Event end date accuracy: Uncertainty or accuracy of event end date."],
            [
                "Is housing destruction: This field  indicates whether the displacement data includes individuals "
                "displaced by housing destruction. Values are \"Yes\" if the data reflects households whose homes were "
                "destroyed and \"No\" otherwise. This field relies on the data specified in \"Reported Figures\" and "
                "is linked to the \"Unit\" of measurement used, which in this context refers to houses destroyed."
            ],
            [
                "Violence type: This field categorizes the type of violence using IDMC's typology, which aligns with "
                "international classifications. The categories include\n"
                "- International Armed Conflict (IAC): Refers to armed conflict between two or more states.\n"
                "- Non-International Armed Conflict (NIAC): Refers to protracted armed conflict occurring within the "
                "territory of a single state between its government and non-state armed groups, or between such groups "
                "themselves.\n"
                "- Unclear/Unknown: Indicates situations where the type of violence is not definitively categorized "
                "due to limited information."
            ],
            [
                "Event codes (Code:Type): Unique codes such as the GLIDE number and other database-specific codes used "
                "to identify and track specific events across various databases."
            ],
            [
                "Locations name: This field indicates the names of locations where displacement incidents have been "
                "reported. It's important to note that this field may exhibit a many-to-one relationship, signifying "
                "that multiple location names could be associated with a single reported figure, preventing "
                "disaggregation by individual location. This becomes particularly relevant in geospatial analysis, "
                "where Geographic Information System (GIS) software may interpret these multi-point entities as single "
                "data points, potentially leading to the inadvertent double-counting of figures. To mitigate this "
                "issue, it's advisable to preprocess the dataset by either dividing the total figure by the number of "
                "locations or distributing the \"Total figures\" values based on a weighting factor such as population "
                "density. This ensures a more accurate representation of the displacement data across individual "
                "locations and prevents duplication of figures during analysis."
            ],
            [
                "Locations coordinates: This field contains geographic coordinates representing the reported locations. "
                "Please note that this field contains multipoints  meaning that multiple locations may represent one "
                "figures. It's important to note that this field may exhibit a many-to-one relationship, signifying "
                "that multiple location names could be associated with a single reported figure, preventing "
                "disaggregation by individual location. This becomes particularly relevant in geospatial analysis, "
                "where Geographic Information System (GIS) software may interpret these multi-point entities as single "
                "data points, potentially leading to the inadvertent double-counting of figures. To mitigate this "
                "issue, it's advisable to preprocess the dataset by either dividing the total figure by the number of "
                "locations or distributing the \"Total figures\" values based on a weighting factor such as population "
                "density. This ensures a more accurate representation of the displacement data across individual "
                "locations and prevents duplication of figures during analysis."
            ],
            [
                "Locations accuracy:  This field indicates the estimated precision of the reported locations. It "
                "serves as a clue to the likely administrative unit level (e.g., country, state, district) used for "
                "reporting. "
            ],
            [
                "Locations type: This field specifies the type of displacement location within a reported event. It "
                "can indicate, Origin: The place where people were displaced from. Destination: The location where "
                "displaced people arrived. Both: In some cases, both origin and destination information might be "
                "included. It's crucial to note that different locations reported for a single figure may pertain to "
                "both the origin and destination of displacement incidents. This distinction is particularly salient "
                "in geospatial analysis, where Geographic Information System (GIS) software may interpret these "
                "multi-point entities as singular data points, potentially resulting in inadvertent double-counting of "
                "figures. To mitigate this issue, it is recommended to preprocess the dataset prior to GIS analysis to "
                "ensure accurate representation and avoid duplication of figures."
            ],
            [
                "Displacement occurred: This field contains values that represent if preventive evacuations were "
                "reported. These evacuations are the result of existing early warning systems."
            ],
        ]
        for item in data_description_1:
            ws3.append(item)
        ws3.append([])
        ws3.append(['DATA DESCRIPTION: 2_Context_Displacement_data table'])
        ws3.append([])
        ws3.append([
            'This dataset provides contextual information and analysis documented by IDMC analysts. It captures flags '
            'related to methodology, caveats, sources, and challenges identified for each metric, reporting year, and '
            'country.'
        ])
        ws3.append([])

        data_description_2 = [
            ["ISO3: Represents the ISO 3166-1 alpha-3 code, the code 'AB9' is assigned to the Abyei Area."],
            ["Year: Indicates the year for which displacement data are reported."],
            ["Figure_Cause_Name: Identifies the trigger of displacement, such as conflict or disasters."],
            [
                "Figure_Category_Name: Categorizes the type of displacement metric. It details values for Internal "
                "Displacements (internal displacement flows) and Total Number of IDPs (internal displacement stocks), "
                "as defined earlier in this document."
            ],
            [
                "Description: Provides contextual information about the data, including sources and data limitations. "
                "It is essential for representing the analysis conducted by IDMC analysts. This field also details the "
                "methodology used, descriptions of sources, and outlines any caveats and challenges identified with "
                "the displacement figures reported."
            ],
            [
                "Figures: Represents the total number of internal displacements or IDPs. For internal displacements, "
                "units are recorded as 'internal displacement flows' or 'internal displacement movements.' For total "
                "number of IDPs, units reflect the total number of people living in displacement."
            ],
            [
                "Figures_Rounded: Displays rounded figures to provide a simplified view of the data that matches the "
                "figures reported in the Global Report on Internal Displacement (GRID)."
            ],
        ]
        for item in data_description_2:
            ws3.append(item)

        qs = qs.filter(
            locations_coordinates__isnull=False,
        ).annotate(
            event_main_trigger=Case(
                When(
                    gidd_event__cause=Crisis.CRISIS_TYPE.CONFLICT,
                    then=F('gidd_event__violence__name')
                ),
                When(
                    gidd_event__cause=Crisis.CRISIS_TYPE.DISASTER,
                    then=F('gidd_event__disaster_sub_type__name')
                ),
                When(
                    gidd_event__cause=Crisis.CRISIS_TYPE.OTHER,
                    then=F('gidd_event__other_sub_type__name')
                ),
                output_field=models.CharField(),
            ),
        )

        for item in qs:
            ws.append([
                item.figure_id,
                item.iso3,
                item.country_name,
                item.geographical_region_name,
                self._get_cause(item.cause),
                item.year,
                self._get_category(item.category),
                self._get_unit(item.unit),
                item.reported,
                item.household_size,
                item.total_figures,
                item.disaster_category_name,
                item.disaster_sub_category_name,
                item.disaster_type_name,
                item.disaster_sub_type_name,
                item.violence_name,
                item.other_sub_type_name,
                item.start_date,
                self._get_date_accuracy(item.start_date_accuracy),
                item.end_date,
                self._get_date_accuracy(item.end_date_accuracy),
                item.stock_date,
                self._get_date_accuracy(item.stock_date_accuracy),
                item.stock_reporting_date,
                string_join(EXTERNAL_ARRAY_SEPARATOR, item.publishers),
                string_join(EXTERNAL_ARRAY_SEPARATOR, item.sources),
                string_join(EXTERNAL_ARRAY_SEPARATOR, item.sources_type),
                item.gidd_event.event_id,
                item.gidd_event.name,
                self._get_cause(item.gidd_event.cause),
                item.event_main_trigger,
                item.gidd_event.start_date,
                item.gidd_event.end_date,
                self._get_date_accuracy(item.gidd_event.start_date_accuracy),
                self._get_date_accuracy(item.gidd_event.end_date_accuracy),
                "Yes" if item.is_housing_destruction else "No",
                self.extract_event_data(
                    item.gidd_event.event_codes,
                    item.gidd_event.event_codes_type,
                    item.gidd_event.event_codes_iso3,
                    item.iso3,
                ),
                string_join(EXTERNAL_ARRAY_SEPARATOR, item.locations_coordinates),
                string_join(EXTERNAL_ARRAY_SEPARATOR, item.locations_names),
                _get_location_accuracy_labels(item.locations_accuracy),
                _get_location_type_labels(item.locations_type),
                self._get_displacement_occurred(item.displacement_occurred),
            ])

        response = HttpResponse(content=save_virtual_workbook(wb))
        response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
        response['Content-Type'] = 'application/octet-stream'
        return response

    @extend_schema(responses=DisaggregationSerializer(many=True))
    @action(
        detail=False,
        methods=["get"],
        url_path="disaggregated-geojson",
        permission_classes=[AllowAny],
        pagination_class=None,
    )
    def export_disaggregated_geojson(self, request):
        """
        Export the disaggregated data in geojson format file
        """
        track_gidd(
            self.request.GET.get('client_id'),
            ExternalApiDump.ExternalApiType.GIDD_DISAGGREGATION_EXPORT_GEOJSON,
            viewset=self,
        )
        queryset = GiddFigure.objects.select_related('gidd_event').order_by(
            '-year',
            'iso3',
            'id',
        ).filter(
            year__gte=2023
        )
        qs = self.filter_queryset(queryset)
        return self._export_disaggregated_geojson(qs)

    @extend_schema(responses=DisaggregationSerializer(many=True))
    @action(
        detail=False,
        methods=["get"],
        url_path="disaggregated-export",
        permission_classes=[AllowAny],
        pagination_class=None,
    )
    def export_disaggregated(self, request):
        """
        Export the disaggregated data in excel format file
        """
        track_gidd(
            self.request.GET.get('client_id'),
            ExternalApiDump.ExternalApiType.GIDD_DISAGGREGATION_EXPORT_EXCEL,
            viewset=self,
        )
        queryset = GiddFigure.objects.select_related('gidd_event').order_by(
            '-year',
            'iso3',
            'id',
        ).filter(
            year__gte=2023
        )
        qs = self.filter_queryset(queryset)
        return self._export_disaggregated_excel(qs)


class PublicFigureAnalysisViewSet(ListOnlyViewSetMixin):
    serializer_class = PublicFigureAnalysisSerializer
    filter_backends = (DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter)
    filterset_class = PublicFigureAnalysisFilterSet

    def get_queryset(self):
        track_gidd(
            self.request.GET.get('client_id'),
            ExternalApiDump.ExternalApiType.GIDD_PUBLIC_FIGURE_ANALYSIS_REST,
            viewset=self,
        )
        return PublicFigureAnalysis.objects.all()

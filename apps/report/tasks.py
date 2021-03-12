import logging
from tempfile import NamedTemporaryFile

import dramatiq
from django.core.files.base import ContentFile
from openpyxl import Workbook

from helix.settings import QueuePriority

REPORT_TIMEOUT = 25 * 60 * 1000  # 25 minutes


# @dramatiq.actor(queue_name=QueuePriority.HEAVY.value, max_retries=3, time_limit=REPORT_TIMEOUT)
def generate_report_excel(generation_id):
    from apps.report.models import ReportGeneration
    generation = ReportGeneration.objects.get(id=generation_id)
    generation.status = ReportGeneration.REPORT_GENERATION_STATUS.IN_PROGRESS
    generation.save(update_fields=['status'])
    wb = Workbook()

    try:
        for sheet_name, sheet_data in generation.get_excel_sheets_data().items():
            headers = sheet_data['headers']
            data = sheet_data['data']
            formulae = sheet_data['formulae']
            aggregation = sheet_data.get('aggregation', None)

            ws = wb.create_sheet(sheet_name)
            # primary headers and data
            for idx, (header_key, header_val) in enumerate(headers.items()):
                ws.cell(column=idx + 1, row=1, value=header_val)
                for idy, datum in enumerate(data):
                    ws.cell(column=idx + 1, row=idy + 2, value=datum[header_key])
            # secondary headers and data
            idx2 = 0
            for idx2, (header_key, formula) in enumerate(formulae.items()):
                # column starts at 1, hence idx+idx2+2
                ws.cell(column=idx + idx2 + 2, row=1, value=header_key)
                # list indexing starts at 0, hence idx+idx2+1
                for row, cell in enumerate(list(ws.columns)[idx + idx2 + 1], 1):
                    if row == 1:
                        continue
                    cell.value = formula.format(row=row)
            # add a gap
            column_at = idx + idx2 + 2
            ws.cell(column=column_at, row=1, value='')

            if not aggregation:
                continue
            agg_headers = aggregation['headers']
            agg_data = aggregation['data']
            agg_formulae = aggregation['formulae']

            # primary headers and data
            for idx, (header_key, header_val) in enumerate(agg_headers.items()):
                ws.cell(column=column_at + idx + 1, row=1, value=header_val)
                for idy, datum in enumerate(agg_data):
                    ws.cell(column=column_at + idx + 1, row=idy + 2, value=datum[header_key])
            # secondary headers and data
            for idx2, (header_key, formula) in enumerate(agg_formulae.items()):
                # column starts at 1, hence idx+idx2+2
                ws.cell(column=column_at + idx + idx2 + 2, row=1, value=header_key)
                # list indexing starts at 0, hence idx+idx2+1
                for row, cell in enumerate(list(ws.columns)[column_at + idx + idx2 + 1], 1):
                    if row == 1:
                        continue
                    cell.value = formula.format(row=row)

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            content = tmp.read()
            path = f'{ReportGeneration.FULL_REPORT_FOLDER}/{generation.report.name}.xlsx'
            generation.full_report.save(path, ContentFile(content))
        generation.status = ReportGeneration.REPORT_GENERATION_STATUS.COMPLETED
    except:  # NOQA E722
        generation.status = ReportGeneration.REPORT_GENERATION_STATUS.FAILED

    generation.save(update_fields=['status'])
